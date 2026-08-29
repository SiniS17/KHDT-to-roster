"""
main.py
================
Reads a "KHDT" (training-plan) workbook and a "Roster" workbook, matches
people by Employee ID (cross-checked against Name), and marks the matching
day-columns in the Roster with a training code for every day that falls
inside that person's training date range.

--------------------------------------------------------------------------
HOW TO RUN IT
--------------------------------------------------------------------------
Simplest usage: put this script (or the .exe built from it) in a folder,
create a subfolder called "INPUT" next to it, and drop the two Excel files
in there - one with "KHDT" somewhere in its filename, one with "Roster"
somewhere in its filename (case doesn't matter). Then just run it with no
arguments:

    my_folder/
        main.py            (or main.exe once packaged)
        INPUT/
            KHDT_T9_2026.xlsx
            Roster_Doi_3_thang_9_nam_2026.xlsx

It will create an "OUTPUT" folder next to the script/exe containing the
updated roster (named "<original roster name>_updated.xlsx") plus a
".log.xlsx" workbook with the full run log. A summary also prints to the
console / pops up in a message box.

If there's no INPUT/ folder, it falls back to a simple file-picker dialog,
or you can still pass paths explicitly:

    python main.py path\\to\\KHDT.xlsx path\\to\\Roster.xlsx [--force] [-o out.xlsx]

Designed to be frozen into a standalone .exe with PyInstaller, e.g.:

    pyinstaller --onefile main.py

(Leave off --noconsole if you want the console window to stay open and
show progress/errors as it runs - the script pauses for Enter at the end
when there's no console-less GUI available anyway.)

Only dependency: openpyxl (pip install openpyxl).

--------------------------------------------------------------------------
HOW MATCHING WORKS
--------------------------------------------------------------------------
1. The KHDT sheet is scanned for its header row (it looks for the Vietnamese
   column headers "Mã NV" / "Họ tên" / "Bắt đầu" / "Kết thúc"), so it does
   not depend on a fixed row number.
2. The Roster sheet is scanned for ITS header row the same way (it looks
   for "ID" and "NAME" - case-insensitive - plus a run of real date cells
   for the day columns).
3. For every KHDT row, the Employee ID is looked up in the Roster's ID
   column. If found, the Name is compared too (accent/case/space
   insensitive) just as a sanity check - a mismatch is logged as a warning
   but the row is still processed, since ID is authoritative.
4. For every day between "Bắt đầu" and "Kết thúc" (inclusive) that has a
   matching date column in the Roster, the cell is marked. If several
   courses land on the same person's day, their markers are combined, with
   SÁNG before CHIỀU (e.g. "H VHNT SÁNG-VHDN CHIỀU"). EXCEPTION: if the
   "Lí do" note names one specific date (e.g. "Lên lớp chiều 8/9/2026"),
   only that single day is marked, even if Bắt đầu/Kết thúc span a wider
   range.
5. Saturday and Sunday course assignments are written as a black "N"
   instead of a course marker. Weekday course markers are written in red.
   An existing black "N" is overwritten and logged as an "Overwrite day off";
   an existing red "N" is logged as a "Guarantee day off conflict".

--------------------------------------------------------------------------
DATE BUG WORKAROUND
--------------------------------------------------------------------------
KHDT files have been observed to mix real Excel date values with plain
text dates, which causes Excel/openpyxl to sometimes read a date like
"11/09/2026" (11 Sep) as 9 Nov instead (US month/day order applied to a
value that was typed in day/month order). Since we know which month the
Roster actually covers, dates from KHDT whose month falls outside a small
window around the Roster's month are automatically flipped back
(day<->month) if that produces a plausible date. Anything still
implausible is logged rather than guessed at silently.

--------------------------------------------------------------------------
CUSTOMISING THE MARKER TEXT / ABBREVIATIONS
--------------------------------------------------------------------------
Edit config.py to change abbreviations or the names of the Excel log columns.
"""

import argparse
import re
import sys
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from config import (
    COURSE_ABBREVIATIONS,
    COURSE_NAME_ABBREVIATIONS,
    LOG_COLUMN_HEADERS,
    LOG_EVENT_LABELS,
    MARKER_PREFIX,
)

# ==========================================================================
# CONFIG — safe to edit
# ==========================================================================

# If a cell in the Roster already has content, don't overwrite by default.
DEFAULT_FORCE_OVERWRITE = False

# Font colors used for values this tool writes (ARGB hex).
MARKER_FONT_COLOR = "FFFF0000"  # red
WEEKEND_FONT_COLOR = "FF000000"  # black

# How many days on either side of "today's window" we tolerate before we
# suspect a day/month swap in a KHDT date. (see fix_date_month_swap)
MONTH_SWAP_TOLERANCE = 1

# ==========================================================================
# Helpers
# ==========================================================================


def strip_accents(s: str) -> str:
    if not s:
        return ""
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def normalize_name(s):
    if not s:
        return ""
    s = strip_accents(str(s)).lower()
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_id(s):
    if s is None:
        return ""
    return str(s).strip().upper()


def parse_khdt_date(value):
    """
    Parse a KHDT date cell into a date object (or None).

    KHDT's date columns are stored in YYYY-DD-MM order (day and month
    swapped from the usual YYYY-MM-DD/DD-MM-YYYY convention). This shows up
    two ways in the workbook:
      - Real Excel date/datetime cells: openpyxl hands back Python
        year/month/day fields, but because of how the source data was typed
        in, what Python calls ".month" is actually the day and what it
        calls ".day" is actually the month. We swap them back here.
      - Plain text cells: written as "DD/MM/YYYY" (normal order, e.g.
        "10/09/2026") or occasionally "YYYY-DD-MM" (e.g. "2026-11-09" for
        9-Nov... i.e. actually day=11, month=09). Both are handled below.
    """
    if value is None:
        return None

    if isinstance(value, (datetime, date)):
        d = value.date() if isinstance(value, datetime) else value
        # Stored as YYYY-DD-MM -> swap back to a real YYYY-MM-DD date.
        try:
            return date(d.year, d.day, d.month)
        except ValueError:
            # Field values don't form a valid swapped date (e.g. "day" > 12)
            # -> assume this cell was already a normal, correct date.
            return d

    if isinstance(value, str):
        v = value.strip()
        if not v:
            return None
        # Slash-separated text is written in normal DD/MM/YYYY order.
        for fmt in ("%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                continue
        # Hyphenated text follows the file's YYYY-DD-MM convention.
        for fmt in ("%Y-%d-%m",):
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                continue
    return None


def check_date_plausible(d, expected_year, expected_month, log, context=""):
    """Log a warning (does not alter the date) if it lands outside the
    Roster's own month/year by more than MONTH_SWAP_TOLERANCE - a sign the
    YYYY-DD-MM fix above didn't apply cleanly to this particular cell."""
    if d is None:
        return
    if d.year == expected_year and abs((d.month - expected_month) % 12) <= MONTH_SWAP_TOLERANCE:
        return
    log(f"[CHECK] {context}: date {d} looks outside the expected {expected_year}-{expected_month:02d} window - please verify")


def extract_note_date(note, expected_year):
    """
    Some KHDT rows have a broad start/end range (e.g. a week of self-study),
    but the "Lí do" note pins down the ONE actual class day, phrased as
    "Lên lớp chiều 8/9/2026" or "Lên lớp sáng 03/9/2026" ("attend class in
    the afternoon/morning of <date>"). When that phrasing is present, that
    single day should be marked instead of the whole range.

    Deliberately narrow: only triggers right after "Lên lớp", so it does
    NOT fire on notes that merely mention some other date in passing (a
    conflicting class's schedule, a "registered but missing from Bravo"
    note, etc.) - those keep the full Bắt đầu/Kết thúc range.
    Returns a date, or None if no such pinned date is found.
    """
    if not note:
        return None
    text = str(note)
    m = re.search(r"lên\s*lớp.{0,20}?(\d{1,2})/(\d{1,2})/(\d{2,4})", text, re.IGNORECASE)
    if not m:
        return None
    day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if year < 100:
        year += 2000
    try:
        return date(year, month, day)
    except ValueError:
        return None


def set_cell_value(cell, value, color):
    """Write a value while preserving the cell's other font attributes."""
    cell.value = value
    f = cell.font
    cell.font = Font(
        name=f.name,
        size=f.size,
        bold=f.bold,
        italic=f.italic,
        vertAlign=f.vertAlign,
        underline=f.underline,
        strike=f.strike,
        color=color,
    )


def set_marker(cell, marker):
    """Write a weekday course marker in red."""
    set_cell_value(cell, marker, MARKER_FONT_COLOR)


def set_weekend_marker(cell):
    """Write the weekend non-working marker in black."""
    set_cell_value(cell, "N", WEEKEND_FONT_COLOR)


def abbreviate_course_name(course_name):
    """
    Return a compact label for a course name when no class ID is available.

    Known long names use explicit, readable labels. Unknown names fall back
    to initials of meaningful words, keeping short codes and words containing
    digits intact.
    """
    if not course_name:
        return None

    first_line = " ".join(str(course_name).strip().splitlines()[0].split())
    if not first_line:
        return None

    normalized_name = normalize_name(first_line)
    for full_name, abbreviation in COURSE_NAME_ABBREVIATIONS.items():
        if normalize_name(full_name) == normalized_name:
            return abbreviation

    stop_words = {
        "a", "an", "and", "for", "in", "of", "on", "part", "the", "to",
        "va", "ve", "và", "về", "của", "cho", "trong",
    }
    stop_words = {normalize_name(word) for word in stop_words}
    tokens = re.findall(r"[A-Za-zÀ-ỹ0-9]+(?:[/+.-][A-Za-zÀ-ỹ0-9]+)*", first_line)
    meaningful = [token for token in tokens if normalize_name(token) not in stop_words]
    if not meaningful:
        return first_line[:24]

    pieces = []
    for token in meaningful[:8]:
        normalized_token = strip_accents(token)
        if any(character.isdigit() for character in token) or token.isupper() or len(token) <= 4:
            pieces.append(normalized_token.upper())
        else:
            pieces.append(normalized_token[0].upper())
    return " ".join(pieces)[:24].rstrip()


def is_course_marker(value):
    """Return whether a cell contains one of this tool's course markers."""
    if not isinstance(value, str):
        return False
    return value.strip().upper().startswith(f"{MARKER_PREFIX.upper()} ")


def is_n_value(value):
    """Return whether a cell contains exactly the roster day-off marker."""
    return isinstance(value, str) and value.strip().upper() == "N"


def _font_color_matches(cell, target_color, default_is_match=False):
    """Match the common RGB/theme forms used by openpyxl for font colors."""
    color = cell.font.color
    if color is None:
        return default_is_match

    target_rgb = target_color[-6:].upper()
    if color.type == "rgb":
        return bool(color.rgb) and color.rgb[-6:].upper() == target_rgb
    if color.type == "theme":
        # Excel theme 1 is Dark 1, the default black text color.
        return target_rgb == "000000" and color.theme == 1
    if color.type == "indexed":
        # Indexed color 8 is black in the standard Excel palette.
        return target_rgb == "000000" and color.indexed == 8
    return False


def is_black_n(cell):
    """Return whether the cell is an exact black day-off marker."""
    return is_n_value(cell.value) and _font_color_matches(
        cell, WEEKEND_FONT_COLOR, default_is_match=True
    )


def is_red_n(cell):
    """Return whether the cell is an exact red guaranteed-day-off marker."""
    return is_n_value(cell.value) and _font_color_matches(cell, MARKER_FONT_COLOR)


def append_to_existing(existing, addition):
    """Preserve existing roster content while adding a new course/value."""
    if existing in (None, ""):
        return addition
    if is_course_marker(existing):
        return combine_markers([existing, addition])
    return f"{str(existing).strip()}-{addition}"


def _marker_sort_key(marker):
    """Sort morning before afternoon while keeping other markers last."""
    normalized = normalize_name(marker)
    if "sang" in normalized:
        return 0
    if "chieu" in normalized:
        return 1
    return 2


def combine_markers(markers):
    """
    Combine course markers without repeating the common "H" prefix.

    Markers are deduplicated and ordered by session so that a morning and an
    afternoon course are easy to read in one cell.
    """
    unique = []
    seen = set()
    for marker in markers:
        if not marker:
            continue
        marker = re.sub(r"\s+", " ", str(marker).strip())
        key = marker.casefold()
        if key not in seen:
            seen.add(key)
            unique.append(marker)

    unique.sort(key=_marker_sort_key)
    if not unique:
        return ""

    prefix = f"{MARKER_PREFIX} "
    parts = [
        marker[len(prefix):].strip() if marker.upper().startswith(prefix.upper()) else marker
        for marker in unique
    ]
    return f"{MARKER_PREFIX} {'-'.join(parts)}".strip()


def record_log_event(event_log, name, employee_id, event, reason, event_date=None, details=None):
    """Aggregate issue details by employee and date for the Excel log."""
    if event_log is None:
        return
    if isinstance(event_date, datetime):
        event_date = event_date.date()

    name = str(name or "").strip()
    employee_id = normalize_id(employee_id)
    key = (name, employee_id, event_date)
    record = event_log.setdefault(
        key,
        {
            "name": name,
            "id": employee_id,
            "date": event_date,
            "events": [],
            "reasons": [],
            "details": [],
        },
    )
    if event and event not in record["events"]:
        record["events"].append(event)
    if reason and reason not in record["reasons"]:
        record["reasons"].append(str(reason))
    if details and details not in record["details"]:
        record["details"].append(str(details))


def find_khdt_header(ws):
    """Locate the KHDT header row and return a dict of column indices."""
    wanted = {
        "name": ("họ tên", "ho ten"),
        "id": ("mã nv", "ma nv"),
        "doi": ("phòng/đội", "phong/doi"),
        "course": ("tên kđt", "ten kdt"),
        "class_code": ("mã lớp học", "ma lop hoc"),
        "start": ("bắt đầu", "bat dau"),
        "end": ("kết thúc", "ket thuc"),
        "note": ("lí do", "lý do", "ly do"),
    }
    for r in range(1, min(ws.max_row, 30) + 1):
        row_vals = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                row_vals[c] = normalize_name(v)
        if not row_vals:
            continue
        found = {}
        for key, aliases in wanted.items():
            norm_aliases = [normalize_name(a) for a in aliases]
            for c, norm in row_vals.items():
                if any(a in norm for a in norm_aliases):
                    found[key] = c
                    break
        if "id" in found and "name" in found and "start" in found:
            return r, found
    raise ValueError("Could not find the KHDT header row (looked for 'Mã NV' / 'Họ tên' / 'Bắt đầu').")


def find_roster_header(ws):
    """
    Locate the Roster header row (has 'ID' and 'NAME' columns) and the set
    of day-columns (a run of real date cells) on that same row.
    Returns (header_row, id_col, name_col, {date: col_index, ...})
    """
    for r in range(1, min(ws.max_row, 30) + 1):
        id_col = None
        name_col = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            norm = normalize_name(v)
            if norm == "id":
                id_col = c
            elif "name" in norm and "surname" in norm or norm == "name":
                name_col = c
        if id_col and name_col:
            date_cols = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, (datetime, date)):
                    d = v.date() if isinstance(v, datetime) else v
                    date_cols[d] = c
            if date_cols:
                return r, id_col, name_col, date_cols
    raise ValueError("Could not find the Roster header row (looked for 'ID' + 'NAME' columns plus date columns).")


def build_marker(class_code, course_name, note):
    label = None
    if class_code:
        code = str(class_code)
        # Strip the trailing "/<batch>/<S>" segments if present.
        parts = code.rsplit("/", 2)
        base = parts[0] if len(parts) >= 2 else code
        label = COURSE_ABBREVIATIONS.get(base, base)
    elif course_name:
        label = abbreviate_course_name(course_name)

    if not label:
        label = "?"

    marker = f"{MARKER_PREFIX} {label}".strip()

    note_l = normalize_name(note) if note else ""
    if "sang" in note_l and "sang" not in normalize_name(marker):
        marker += " SÁNG"
    elif "chieu" in note_l or "chiều" in (note or "").lower():
        marker += " CHIỀU"

    return marker


# ==========================================================================
# Core
# ==========================================================================


def load_khdt(khdt_path):
    wb = openpyxl.load_workbook(khdt_path, data_only=True)
    ws = wb.worksheets[0]
    header_row, cols = find_khdt_header(ws)

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        emp_id = ws.cell(row=r, column=cols["id"]).value
        if not emp_id:
            continue
        name = ws.cell(row=r, column=cols["name"]).value
        doi = ws.cell(row=r, column=cols.get("doi", 0)).value if cols.get("doi") else None
        course = ws.cell(row=r, column=cols.get("course", 0)).value if cols.get("course") else None
        class_code = ws.cell(row=r, column=cols.get("class_code", 0)).value if cols.get("class_code") else None
        start_raw = ws.cell(row=r, column=cols["start"]).value
        end_raw = ws.cell(row=r, column=cols.get("end", cols["start"])).value if cols.get("end") else start_raw
        note = ws.cell(row=r, column=cols.get("note", 0)).value if cols.get("note") else None

        start = parse_khdt_date(start_raw)
        end = parse_khdt_date(end_raw)
        if end is None:
            end = start

        rows.append(
            {
                "row": r,
                "id": normalize_id(emp_id),
                "name": name,
                "doi": doi,
                "course": course,
                "class_code": class_code,
                "start": start,
                "end": end,
                "note": note,
            }
        )
    return rows


def mark_roster(khdt_rows, roster_path, output_path, force=False, log=print, event_log=None):
    wb = openpyxl.load_workbook(roster_path)
    ws = wb.worksheets[0]
    header_row, id_col, name_col, date_cols = find_roster_header(ws)

    if date_cols:
        sample_dates = sorted(date_cols.keys())
        expected_year = sample_dates[0].year
        # Most common month among the day columns
        months = [d.month for d in sample_dates]
        expected_month = max(set(months), key=months.count)
    else:
        raise ValueError("No date columns found in Roster header row.")

    # Build ID -> roster row lookup
    id_to_row = {}
    id_to_name = {}
    for r in range(header_row + 1, ws.max_row + 1):
        rid = ws.cell(row=r, column=id_col).value
        if not rid:
            continue
        norm = normalize_id(rid)
        id_to_row[norm] = r
        id_to_name[norm] = ws.cell(row=r, column=name_col).value

    stats = {
        "marked": 0,
        "conflicts": 0,
        "overlaps_combined": 0,
        "weekend_n": 0,
        "no_id_match": 0,
        "no_date_overlap": 0,
        "name_mismatch": 0,
    }
    assignments = {}

    for entry in khdt_rows:
        eid = entry["id"]
        if eid not in id_to_row:
            stats["no_id_match"] += 1
            continue

        roster_row = id_to_row[eid]
        roster_name = id_to_name.get(eid)
        if normalize_name(entry["name"]) and normalize_name(entry["name"]) != normalize_name(roster_name):
            stats["name_mismatch"] += 1
            record_log_event(
                event_log,
                entry["name"],
                eid,
                LOG_EVENT_LABELS["name_mismatch"],
                f"KHDT name: {entry['name']}; Roster name: {roster_name}",
            )
            log(
                f"[WARN] ID {eid} matched but names differ: "
                f"KHDT='{entry['name']}' vs Roster='{roster_name}' (row {entry['row']})"
            )

        start, end = entry["start"], entry["end"]
        if start is None:
            record_log_event(
                event_log,
                entry["name"],
                eid,
                LOG_EVENT_LABELS["skipped"],
                "No usable start date",
                details=f"KHDT row {entry['row']}",
            )
            log(f"[SKIP] {entry['name']} ({eid}): no usable start date (KHDT row {entry['row']})")
            continue
        if end is None or end < start:
            end = start
        ctx = f"{entry['name']} ({eid}), KHDT row {entry['row']}"
        check_date_plausible(start, expected_year, expected_month, log, ctx + " [start]")
        check_date_plausible(end, expected_year, expected_month, log, ctx + " [end]")

        note_date = extract_note_date(entry["note"], expected_year)
        if note_date is not None and (note_date != start or note_date != end):
            record_log_event(
                event_log,
                entry["name"],
                eid,
                LOG_EVENT_LABELS["note"],
                str(entry["note"]).strip(),
                event_date=note_date,
                details=f"Date was pinned to {note_date} instead of the range {start}..{end}",
            )
            log(
                f"[NOTE DATE] {entry['name']} ({eid}): range was {start}..{end}, "
                f"but Lí do pins it to {note_date} -> marking that day only (KHDT row {entry['row']})"
            )
            start = end = note_date

        marker = build_marker(entry["class_code"], entry["course"], entry["note"])

        d = start
        matched_any_day = False
        while d <= end:
            if d in date_cols:
                matched_any_day = True
                col = date_cols[d]
                key = (roster_row, d)
                assignments.setdefault(key, []).append(
                    {
                        "marker": marker,
                        "name": entry["name"],
                        "id": eid,
                        "row": entry["row"],
                        "column": col,
                    }
                )
            d += timedelta(days=1)

        if not matched_any_day:
            stats["no_date_overlap"] += 1
            record_log_event(
                event_log,
                entry["name"],
                eid,
                LOG_EVENT_LABELS["no_date_overlap"],
                f"{start}..{end} does not fall in the roster's date range",
                details=f"KHDT row {entry['row']}",
            )
            log(
                f"[NO OVERLAP] {entry['name']} ({eid}): {start}..{end} doesn't fall in this "
                f"Roster's date range (KHDT row {entry['row']})"
            )

    for (roster_row, d), entries in assignments.items():
        col = entries[0]["column"]
        cell = ws.cell(row=roster_row, column=col)
        existing = cell.value
        markers = [item["marker"] for item in entries]
        combined = combine_markers(markers)

        if len(set(marker.casefold() for marker in markers)) > 1:
            stats["overlaps_combined"] += 1
            record_log_event(
                event_log,
                entries[0]["name"],
                entries[0]["id"],
                LOG_EVENT_LABELS["overlap"],
                f"Courses combined: {', '.join(markers)} -> {combined}",
                event_date=d,
                details=f"KHDT rows: {', '.join(str(item['row']) for item in entries)}",
            )
            log(
                f"[OVERLAP] {entries[0]['name']} ({entries[0]['id']}) {d}: "
                f"combined {', '.join(markers)} -> '{combined}'"
            )

        if d.weekday() >= 5:
            # Weekend cells are always N for course assignments. Existing
            # course markers are replaced. A black N is an overwrite-not-
            # conflict case. Red N and other content are conflicts; the
            # existing content is preserved and N is appended unless --force
            # explicitly requests replacement.
            weekend_value = "N"
            if is_black_n(cell):
                record_log_event(
                    event_log,
                    entries[0]["name"],
                    entries[0]["id"],
                    LOG_EVENT_LABELS["overwrite_day_off"],
                    "Existing black N was overwritten with N",
                    event_date=d,
                    details=f"KHDT row {entries[0]['row']}",
                )
                log(
                    f"[OVERWRITE DAY OFF] {entries[0]['name']} ({entries[0]['id']}) {d}: "
                    "existing black N was overwritten with N "
                    f"(KHDT row {entries[0]['row']})"
                )
            elif existing not in (None, "") and not is_course_marker(existing):
                event_label = (
                    LOG_EVENT_LABELS["guarantee_day_off_conflict"]
                    if is_red_n(cell)
                    else LOG_EVENT_LABELS["conflict"]
                )
                existing_description = (
                    f"Existing red N (guaranteed day off); wanted: N"
                    if is_red_n(cell)
                    else f"Existing cell: {existing}; wanted: N"
                )
                stats["conflicts"] += 1
                record_log_event(
                    event_log,
                    entries[0]["name"],
                    entries[0]["id"],
                    event_label,
                    existing_description,
                    event_date=d,
                    details=f"KHDT row {entries[0]['row']}",
                )
                log(
                    f"[{'GUARANTEE DAY OFF CONFLICT' if is_red_n(cell) else 'CONFLICT'}] "
                    f"{entries[0]['name']} ({entries[0]['id']}) {d}: "
                    f"cell already has '{existing}', wanted 'N'; "
                    f"{'appended N instead' if not force else 'overwrote it'} "
                    f"(KHDT row {entries[0]['row']})"
                )
                if not force:
                    weekend_value = append_to_existing(existing, "N")
            set_cell_value(cell, weekend_value, WEEKEND_FONT_COLOR)
            stats["marked"] += 1
            stats["weekend_n"] += 1
            continue

        combined_with_existing = combined
        if is_black_n(cell):
            record_log_event(
                event_log,
                entries[0]["name"],
                entries[0]["id"],
                LOG_EVENT_LABELS["overwrite_day_off"],
                f"Existing black N was overwritten with {combined}",
                event_date=d,
                details=f"KHDT row {entries[0]['row']}",
            )
            log(
                f"[OVERWRITE DAY OFF] {entries[0]['name']} ({entries[0]['id']}) {d}: "
                f"existing black N was overwritten with '{combined}' "
                f"(KHDT row {entries[0]['row']})"
            )
        elif existing not in (None, ""):
            if is_course_marker(existing):
                combined_with_existing = combine_markers([existing, combined])
            elif is_red_n(cell):
                stats["conflicts"] += 1
                record_log_event(
                    event_log,
                    entries[0]["name"],
                    entries[0]["id"],
                    LOG_EVENT_LABELS["guarantee_day_off_conflict"],
                    f"Existing red N (guaranteed day off); wanted: {combined}",
                    event_date=d,
                    details=f"KHDT row {entries[0]['row']}",
                )
                log(
                    f"[GUARANTEE DAY OFF CONFLICT] {entries[0]['name']} "
                    f"({entries[0]['id']}) {d}: cell already has red N, wanted "
                    f"'{combined}'; "
                    f"{'appended course instead' if not force else 'overwrote it'} "
                    f"(KHDT row {entries[0]['row']})"
                )
                if not force:
                    combined_with_existing = append_to_existing(existing, combined)
            elif not force:
                stats["conflicts"] += 1
                record_log_event(
                    event_log,
                    entries[0]["name"],
                    entries[0]["id"],
                    LOG_EVENT_LABELS["conflict"],
                    f"Existing cell: {existing}; wanted: {combined}",
                    event_date=d,
                    details=f"KHDT row {entries[0]['row']}",
                )
                log(
                    f"[CONFLICT] {entries[0]['name']} ({entries[0]['id']}) {d}: "
                    f"cell already has '{existing}', wanted '{combined}'; appended course instead "
                    f"(KHDT row {entries[0]['row']})"
                )
                combined_with_existing = append_to_existing(existing, combined)

        set_marker(cell, combined_with_existing)
        stats["marked"] += 1

    wb.save(output_path)
    return stats


# ==========================================================================
# CLI / entry point
# ==========================================================================


def get_base_dir():
    """Directory the .exe (or this script) lives in - INPUT/OUTPUT are relative to this."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def find_input_files(input_dir):
    """
    Look inside input_dir for exactly one *.xlsx whose name contains "khdt"
    and exactly one whose name contains "roster" (case-insensitive).
    Returns (khdt_path_or_None, roster_path_or_None, list_of_problem_strings).
    """
    xlsx_files = [p for p in input_dir.glob("*.xlsx") if not p.name.startswith("~$")]
    khdt_matches = [p for p in xlsx_files if "khdt" in p.stem.lower()]
    roster_matches = [p for p in xlsx_files if "roster" in p.stem.lower()]

    problems = []
    khdt_path = None
    roster_path = None

    if len(khdt_matches) == 0:
        problems.append(f'No file with "KHDT" in its name found in {input_dir}')
    elif len(khdt_matches) > 1:
        problems.append("Multiple KHDT-named files found: " + ", ".join(p.name for p in khdt_matches))
    else:
        khdt_path = khdt_matches[0]

    if len(roster_matches) == 0:
        problems.append(f'No file with "Roster" in its name found in {input_dir}')
    elif len(roster_matches) > 1:
        problems.append("Multiple Roster-named files found: " + ", ".join(p.name for p in roster_matches))
    else:
        roster_path = roster_matches[0]

    return khdt_path, roster_path, problems


def write_log_workbook(log_path, event_log, log_lines, stats):
    """Write structured issue details and the raw run log to an XLSX file."""
    wb = Workbook()
    log_ws = wb.active
    log_ws.title = "Log"
    headers = [
        LOG_COLUMN_HEADERS["name"],
        LOG_COLUMN_HEADERS["id"],
        LOG_COLUMN_HEADERS["date"],
        LOG_COLUMN_HEADERS["event"],
        LOG_COLUMN_HEADERS["reason"],
        LOG_COLUMN_HEADERS["details"],
    ]
    log_ws.append(headers)

    records = sorted(
        event_log.values(),
        key=lambda record: (
            record["date"] is None,
            record["date"] or date.max,
            record["id"],
            record["name"],
        ),
    )
    for record in records:
        log_ws.append(
            [
                record["name"],
                record["id"],
                record["date"],
                ", ".join(record["events"]),
                "\n".join(record["reasons"]),
                "\n".join(record["details"]),
            ]
        )

    summary_ws = wb.create_sheet("Run summary")
    summary_ws.append(["Metric", "Value"])
    for metric, value in (
        ("Cells marked", stats["marked"]),
        ("Weekend N cells", stats["weekend_n"]),
        ("Overlaps combined", stats["overlaps_combined"]),
        ("Conflicts noted", stats["conflicts"]),
        ("IDs not in roster", stats["no_id_match"]),
        ("No date overlap", stats["no_date_overlap"]),
        ("Name/ID mismatches", stats["name_mismatch"]),
    ):
        summary_ws.append([metric, value])

    console_ws = wb.create_sheet("Console log")
    console_ws.append(["Message"])
    for line in log_lines:
        console_ws.append([line])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for ws in (log_ws, summary_ws, console_ws):
        for cell in ws[1]:
            cell.font = Font(color="FFFFFFFF", bold=True)
            cell.fill = header_fill
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

    for row in log_ws.iter_rows(min_row=2):
        row[2].number_format = "dd/mm/yyyy"
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    log_ws.auto_filter.ref = log_ws.dimensions
    log_ws.column_dimensions["A"].width = 28
    log_ws.column_dimensions["B"].width = 15
    log_ws.column_dimensions["C"].width = 13
    log_ws.column_dimensions["D"].width = 20
    log_ws.column_dimensions["E"].width = 60
    log_ws.column_dimensions["F"].width = 42

    summary_ws.column_dimensions["A"].width = 25
    summary_ws.column_dimensions["B"].width = 18
    console_ws.column_dimensions["A"].width = 120
    for row in console_ws.iter_rows(min_row=2):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(log_path)


def run(
    khdt_path,
    roster_path,
    output_path=None,
    force=DEFAULT_FORCE_OVERWRITE,
    log=print,
    event_log=None,
):
    khdt_path = Path(khdt_path)
    roster_path = Path(roster_path)
    if output_path is None:
        output_path = roster_path.with_name(roster_path.stem + "_updated" + roster_path.suffix)
    else:
        output_path = Path(output_path)

    log(f"Reading KHDT: {khdt_path}")
    khdt_rows = load_khdt(khdt_path)
    log(f"  -> {len(khdt_rows)} training-assignment rows found")

    log(f"Reading Roster: {roster_path}")
    stats = mark_roster(
        khdt_rows,
        roster_path,
        output_path,
        force=force,
        log=log,
        event_log=event_log,
    )

    log("")
    log("===== Summary =====")
    log(f"Cells marked        : {stats['marked']}")
    log(f"Weekend N cells     : {stats['weekend_n']}")
    log(f"Overlaps combined   : {stats['overlaps_combined']}")
    log(f"Conflicts noted     : {stats['conflicts']}  (use --force to overwrite instead)")
    log(f"ID not in roster    : {stats['no_id_match']}  (not a member of this team, or ID typo)")
    log(f"No date overlap     : {stats['no_date_overlap']}  (training dates outside this roster's month)")
    log(f"Name/ID mismatches  : {stats['name_mismatch']}  (flagged above, still processed by ID)")
    log(f"Saved -> {output_path}")
    return output_path, stats


def main():
    parser = argparse.ArgumentParser(description="Mark a Roster file with training days from a KHDT file.")
    parser.add_argument("khdt", nargs="?", help="Path to the KHDT (training plan) .xlsx file")
    parser.add_argument("roster", nargs="?", help="Path to the Roster .xlsx file")
    parser.add_argument("-o", "--output", help="Output path (default: OUTPUT/<roster>_updated.xlsx)")
    parser.add_argument("--force", action="store_true", help="Overwrite cells that already have content")
    args = parser.parse_args()

    khdt_path, roster_path, output_path = args.khdt, args.roster, args.output

    def have_tk():
        try:
            import tkinter  # noqa: F401
            return True
        except ImportError:
            return False

    def report_fatal(msg):
        print(msg)
        if have_tk():
            import tkinter as tk
            from tkinter import messagebox
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Error", msg)
        else:
            input("Press Enter to exit...")

    # -------------------------------------------------------------
    # 1) Explicit CLI args always win.
    # -------------------------------------------------------------
    if khdt_path and roster_path:
        log_lines = []
        event_log = {}
        out_path, stats = run(
            khdt_path,
            roster_path,
            output_path,
            args.force,
            log=log_lines.append,
            event_log=event_log,
        )
        log_path = Path(out_path).with_suffix(".log.xlsx")
        write_log_workbook(log_path, event_log, log_lines, stats)
        print(f"Saved -> {out_path}")
        print(f"Log -> {log_path}")
        return

    # -------------------------------------------------------------
    # 2) INPUT/ folder next to the .exe (or this script): drop both
    #    files in there, named so one contains "KHDT" and the other
    #    "Roster" (case-insensitive) - no prompts needed.
    # -------------------------------------------------------------
    base_dir = get_base_dir()
    input_dir = base_dir / "INPUT"

    if input_dir.is_dir():
        found_khdt, found_roster, problems = find_input_files(input_dir)
        if problems:
            report_fatal("Couldn't auto-detect the input files in " + str(input_dir) + ":\n- " + "\n- ".join(problems))
            return

        output_dir = base_dir / "OUTPUT"
        output_dir.mkdir(exist_ok=True)
        if not output_path:
            output_path = output_dir / (found_roster.stem + "_updated" + found_roster.suffix)

        log_lines = []
        event_log = {}
        try:
            out_path, stats = run(
                found_khdt,
                found_roster,
                output_path,
                args.force,
                log=log_lines.append,
                event_log=event_log,
            )
        except Exception as e:
            report_fatal(f"{type(e).__name__}: {e}")
            return

        log_path = Path(out_path).with_suffix(".log.xlsx")
        write_log_workbook(log_path, event_log, log_lines, stats)

        summary = (
            f"Marked: {stats['marked']}\n"
            f"Weekend N cells: {stats['weekend_n']}\n"
            f"Overlaps combined: {stats['overlaps_combined']}\n"
            f"Conflicts noted: {stats['conflicts']}\n"
            f"ID not in roster: {stats['no_id_match']}\n"
            f"No date overlap: {stats['no_date_overlap']}\n"
            f"Name/ID mismatches: {stats['name_mismatch']}\n\n"
            f"Saved to:\n{out_path}\n\n"
            f"Full log:\n{log_path}"
        )
        print(summary)
        if have_tk():
            import tkinter as tk
            from tkinter import messagebox
            root = tk.Tk()
            root.withdraw()
            messagebox.showinfo("Done", summary)
        else:
            input("Press Enter to exit...")
        return

    # -------------------------------------------------------------
    # 3) No INPUT/ folder and no CLI args -> fall back to a simple
    #    file-picker so the tool is still usable elsewhere.
    # -------------------------------------------------------------
    if not have_tk():
        parser.error("khdt and roster paths are required (no CLI args, no INPUT/ folder, and no tkinter available)")
        return

    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()
    if not khdt_path:
        khdt_path = filedialog.askopenfilename(title="Select the KHDT (training plan) file", filetypes=[("Excel files", "*.xlsx")])
    if not roster_path:
        roster_path = filedialog.askopenfilename(title="Select the Roster file", filetypes=[("Excel files", "*.xlsx")])
    if not khdt_path or not roster_path:
        return

    log_lines = []
    event_log = {}
    try:
        out_path, stats = run(
            khdt_path,
            roster_path,
            output_path,
            args.force,
            log=log_lines.append,
            event_log=event_log,
        )
    except Exception as e:
        messagebox.showerror("Error", f"{type(e).__name__}: {e}")
        return

    summary = (
        f"Marked: {stats['marked']}\n"
        f"Weekend N cells: {stats['weekend_n']}\n"
        f"Overlaps combined: {stats['overlaps_combined']}\n"
        f"Conflicts noted: {stats['conflicts']}\n"
        f"ID not in roster: {stats['no_id_match']}\n"
        f"No date overlap: {stats['no_date_overlap']}\n"
        f"Name/ID mismatches: {stats['name_mismatch']}\n\n"
        f"Saved to:\n{out_path}"
    )
    messagebox.showinfo("Done", summary)

    log_path = Path(out_path).with_suffix(".log.xlsx")
    write_log_workbook(log_path, event_log, log_lines, stats)


if __name__ == "__main__":
    main()